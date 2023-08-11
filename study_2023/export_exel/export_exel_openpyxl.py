import openpyxl

# Создание нового Excel файла
workbook = openpyxl.Workbook()
sheet = workbook.active

# Запись данных в ячейки (пример)
sheet.cell(row=1, column=1, value='Имя')
sheet.cell(row=1, column=2, value='Возраст')
sheet.cell(row=1, column=3, value='Зарплата')

data = [
    ['Алиса', 25, 50000],
    ['Боб', 30, 60000],
    ['Клара', 22, 45000]
]

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=cell_data)

# Сохранение файла
workbook.save('C:\\Users\\halitsyn.y\\PycharmProjects\\My_study_python\\study_2023\\export_exel.xlsx')
